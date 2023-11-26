import openpyxl

path = "/rf-data/TestDataRegression.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name("ATP")
rows = sheet.max_row
columns = sheet.max_column
print (rows)
print (columns)
headerDetails = ""
for r in range (1, rows + 1):
    for c in range (1, columns+1):
        if r==1:
            if headerDetails == "":
                headerDetails = sheet.cell(row=r, column=c).value
            else:
                headerDetails = headerDetails + ", " + sheet.cell(row=r,column=c).value
        
        testCaseID = sheet.cell(row=r, column=1).value
        if testCaseID == "288768":
            print(sheet.cell(row=r, column=c).value, end="### ")
            