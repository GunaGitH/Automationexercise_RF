from openpyxl import Workbook
from collections import OrderedDict

wb = Workbook()
wb['Sheet'].title='Collections'

od=OrderedDict()
od['Catalog']="50W36"
od['Price']="$100"

sheet = wb.active
columns = sheet.max_column

i=0
for key,value in od.items():
    i=i+1
    for c in range(1, columns+1):
        sheet.cell(row=1, column=i).value = key
        sheet.cell(row=2, column=i).value = value

wb.save("FinalReport.xlsx")