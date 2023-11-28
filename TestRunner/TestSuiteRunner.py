import openpyxl
import os
from pathlib import Path
from FileUtilities import FileUtils
from openpyxl import Workbook
from collections import OrderedDict







scriptLocation = os.getcwd() + "/../Tests/"
scriptToExecute = scriptLocation + "Lennox_1.robot"
reportFolder = scriptLocation + "../Reports/LatestReports"
commandToExecute = "robot -d" + reportFolder + " " + scriptToExecute
print("Executing Test ~"+commandToExecute)
os.system('cmd /c ' + commandToExecute)

dataToSAPLocation = os.getcwd() + "/../rf-data/DataToSAP.txt"

file1 = open(dataToSAPLocation,'r')
Lines = file1.readlines()

count = 0
columnHeader = ""
columnValue = ""


for line in Lines:
    count += 1
    lineDetails = line.strip()
    print("Line{}: {}".format(count,line.strip()))
    if count==1:
        columnHeader = lineDetails
    else:
        columnValue = lineDetails















headerDetails = columnHeader.split("###")
headerValues = columnValue.split("---")
vCount = len(headerValues)
od = OrderedDict()

for i in range(vCount):

    print(str(headerDetails[i])+" = "+str(headerValues[i]))
    od[str(headerDetails[i].strip())] = str(headerValues[i].strip())
    if headerDetails[i].strip().__contains__("Date"):
        print("This is a daate field -> "+headerDetails[i].strip())
        tempDate = headerValues[i].strip()

        if any(chr.isdigit() for chr in tempDate):
            print ("This string contains numbers")
            tmp1 = tempDate.split(",")
            tmp2 = str(tmp1[0].split(" "))
            day = str(tmp2[1])
            month = str(tmp2[0])[0:3]
            print("Formatted -> "+ day + "-" + month)
            od[str(headerDetails[i].strip())] = day + "-" + month
        else:
            print("This string don't contain numbers")



wb = Workbook()
wb['Sheet'].title = 'DataToSAP'

sheet = wb.active
columns = sheet.max_coulmn

i=0
for key,value in od.items():
    i=i+1
    for c in range (i,columns+1):
        sheet.cell(row=1, column=i).value = key
        sheet.cell(row=2, column=i).value = value
        if i>1:
            i=i+2

wb.save(os.getcwd() + "/../rf-data/DataToSAP.xlsx")
print("Successfully Created DataToSAP.xlsx file @ " + os.getcwd() + "/../rf-data/DataToSAP.xlsx")